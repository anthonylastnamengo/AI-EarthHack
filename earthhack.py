import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.cluster import KMeans
from sklearn.metrics.pairwise import cosine_similarity
from textblob import TextBlob
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Load the dataset
file_path = 'AI EarthHack Dataset.csv'  # Replace with your dataset path
data = pd.read_csv(file_path, encoding='ISO-8859-1')

# Ensure all text data is in string format
data['problem'] = data['problem'].astype(str)
data['solution'] = data['solution'].astype(str)

# Text processing and TF-IDF for theme clustering
vectorizer = TfidfVectorizer(stop_words='english', max_features=100)
tfidf_matrix = vectorizer.fit_transform(data['problem'])

# Using KMeans clustering to group ideas into themes
k = 10
model = KMeans(n_clusters=k, n_init=10, random_state=42)
data['theme_cluster'] = model.fit_predict(tfidf_matrix)

# Count the number of ideas in each cluster and identify unique clusters
cluster_counts = data['theme_cluster'].value_counts()
unique_clusters = cluster_counts[cluster_counts < 5].index
data['theme_type'] = data['theme_cluster'].apply(lambda x: 'unique' if x in unique_clusters else 'normal')

# Function to calculate sentiment score
def calculate_sentiment_score(text):
    sentiment = TextBlob(text).sentiment
    tone_score = sentiment.polarity
    length_score = len(text) / data['problem'].str.len().max()
    weighted_score = (length_score * 0.70) + (tone_score * 0.30)
    return weighted_score

# Combine problem and solution into a single text column for analysis
data['combined_text'] = data['problem'] + ' ' + data['solution']

# Calculate sentiment score for each entry
data['sentiment_score'] = data['combined_text'].apply(calculate_sentiment_score)

# Function to calculate uniqueness score
def calculate_uniqueness_score(tfidf_matrix, index):
    cosine_sim = cosine_similarity(tfidf_matrix, tfidf_matrix)
    avg_similarity = sum(cosine_sim[index]) / (len(cosine_sim[index]) - 1)
    uniqueness_score = 1 - avg_similarity
    return uniqueness_score

# Create a column for uniqueness score
data['uniqueness_score'] = 0.0
for cluster in range(k):
    cluster_indices = data[data['theme_cluster'] == cluster].index
    cluster_tfidf_matrix = tfidf_matrix[cluster_indices]
    for index in cluster_indices:
        local_index = list(cluster_indices).index(index)
        data.at[index, 'uniqueness_score'] = calculate_uniqueness_score(cluster_tfidf_matrix, local_index)

# Circular economy vocabulary analysis
ce_keywords = ['sustainability', 'recycle', 'renewable', 'circular economy', 'eco-friendly', 'biodegradable', 
               'carbon footprint', 'renewable energy', 'sustainable development']  # Add up to 100 keywords
def calculate_ce_score(text, keywords):
    score = sum(text.count(word) for word in keywords) / len(keywords)
    return score

# Calculate CE score for each entry
data['ce_score'] = data['combined_text'].apply(lambda x: calculate_ce_score(x, ce_keywords))

# Combine sentiment, uniqueness, and CE scores
data['final_score'] = (data['sentiment_score'] * 1) + (data['uniqueness_score'] * 1) + (data['ce_score'] * 2)

# Apply a multiplier for ideas in the unique theme
data['final_score'] = data.apply(lambda row: row['final_score'] * 1.1 if row['theme_type'] == 'unique' else row['final_score'], axis=1)

# Sort the ideas from best to worst based on the final score
sorted_final_data = data.sort_values(by='final_score', ascending=False)

# Output the top entries based on the final score
sorted_final_data[['id', 'problem', 'solution', 'sentiment_score', 'uniqueness_score', 'ce_score', 'final_score']].to_csv('sorted_ideas.csv', index=False)


# Function to adjust scores based on sentiment score and CE score
def adjust_scores_based_on_sentiment_and_ce(row, sentiment_threshold, ce_threshold):
    adjusted_uniqueness_score = row['uniqueness_score']
    adjusted_ce_score = row['ce_score']
    
    # Adjust uniqueness score if sentiment is high
    if row['sentiment_score'] > sentiment_threshold:
        adjusted_uniqueness_score = row['sentiment_score']
    
    # Adjust CE score if sentiment is high
    if row['sentiment_score'] > sentiment_threshold:
        adjusted_ce_score = row['sentiment_score']

    return adjusted_uniqueness_score, adjusted_ce_score

# Applying the adjustments
sentiment_threshold = 0.8  # Adjust this threshold as needed
ce_threshold = 0.8  # Adjust this threshold as needed
data[['adjusted_uniqueness_score', 'adjusted_ce_score']] = data.apply(lambda row: adjust_scores_based_on_sentiment_and_ce(row, sentiment_threshold, ce_threshold), axis=1, result_type='expand')

# Combine sentiment, adjusted uniqueness, and adjusted CE scores
data['final_score'] = (data['sentiment_score'] * 1) + (data['adjusted_uniqueness_score'] * 1) + (data['adjusted_ce_score'] * 2)

# Apply a multiplier for ideas in the unique theme
data['final_score'] = data.apply(lambda row: row['final_score'] * 1.1 if row['theme_type'] == 'unique' else row['final_score'], axis=1)

# Sort the ideas from best to worst based on the final score
sorted_final_data = data.sort_values(by='final_score', ascending=False)

# Output the top entries based on the final score
sorted_final_data[['id', 'problem', 'solution', 'sentiment_score', 'uniqueness_score', 'ce_score', 'final_score']].to_csv('sorted_ideas_adjusted.csv', index=False)

# Assuming 'sorted_final_data' is your DataFrame with the sorted ideas
# Create a Pandas Excel writer using openpyxl as the engine
excel_writer = pd.ExcelWriter('sorted_ideas_formatted.xlsx', engine='openpyxl')

# Write the sorted DataFrame to an Excel file
sorted_final_data.to_excel(excel_writer, index=False, sheet_name='Sorted Ideas')

# You must save the file before you can use openpyxl to access the workbook
excel_writer._save()

# Now reopen the file using openpyxl to apply the formatting
workbook = openpyxl.load_workbook('sorted_ideas_formatted.xlsx')
worksheet = workbook['Sorted Ideas']

# Apply formatting: Set a fill color for the header row and bold font
header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
bold_font = Font(bold=True)
center_aligned_text = Alignment(horizontal='center')
for cell in worksheet[1]:
    cell.fill = header_fill
    cell.font = bold_font
    cell.alignment = center_aligned_text

# Set the column width to fit content
def as_text(value):
    if value is None:
        return ""
    return str(value)

for col in worksheet.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column name
    for cell in col:
        try:  # Necessary to avoid error on empty cells
            if len(as_text(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    worksheet.column_dimensions[column].width = adjusted_width

# Save the workbook with formatting
workbook.save('sorted_ideas_formatted.xlsx')
